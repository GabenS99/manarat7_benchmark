"""
Output Saver for Prediction Results

Saves prediction results to JSON and Excel formats.
Handles Arabic text encoding and creates comprehensive output files.
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional, List, Set
from datetime import datetime
from collections import defaultdict
import traceback
import re

# Optional YAML support
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

from constants import MCQ_choice_map_en, QuestionType, normalize_model_name, normalize_question_type
from stats_tracker import count_abstentions, is_abstention, calculate_evaluation_statistics

# Check if openpyxl is available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def _generate_filename(
    data: Dict[str, Any],
    original_file_path: Optional[Path] = None
) -> str:
    """
    Generate filename for output files (DRY helper).
    
    Args:
        data: Dataset dictionary with metadata
        original_file_path: Optional path to original dataset file
        
    Returns:
        Filename string (without extension)
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if original_file_path:
        return f"{original_file_path.stem}_{timestamp}_predictions"
    else:
        metadata = data.get("metadata", {})
        source_id = metadata.get("id_source", "dataset")
        source_text = metadata.get("source_text", "dataset")
        difficulty_level = metadata.get("difficulty_level", "dataset")
        question_type = metadata.get("question_type", "dataset")
        return f"{source_id}_{source_text}_{difficulty_level}_{question_type}_{timestamp}_predictions"


def get_model_names(model_dict: Dict[str, Any]) -> List[str]:
    """Get model names from model dictionary."""
    model_names = []
    for provider, model_list in model_dict.items():
        for model_name in model_list:
            model_names.append(f"{provider}/{model_name}")
    return model_names


def extract_prediction_parameters(**kwargs) -> Dict[str, Any]:
    """
    Extract prediction parameters into a compact dict.
    
    This function accepts any prediction parameters as keyword arguments and
    returns a dictionary with only non-None values. This makes it future-proof
    for new parameters without needing to modify the function signature.
    
    Common parameters include:
        - few_shots: Whether to use few-shot examples
        - show_cot: Whether to show chain-of-thought reasoning
        - verbose_instructions: Whether to include detailed instructions
        - abstention: Whether to allow abstention responses
        - verbalized_elicitation: Whether to request confidence scores
        - temperature: Model temperature
        - max_tokens: Maximum tokens for response
        - word_limit: Word limit for response
        - max_retries: Number of retry attempts
        
    Returns:
        Dictionary with prediction parameters (only non-None values included)
    """
    return {k: v for k, v in kwargs.items() if v is not None}


def _build_output_structure(
    data: Dict[str, Any],
    model_dict: Optional[Dict[str, Any]],
    batch_num: Optional[int] = None,
    **prediction_params_kwargs
) -> Dict[str, Any]:
    """
    Build the standard output structure for JSON files (DRY helper).
    
    Args:
        data: Dataset dictionary with metadata and questions
        model_dict: Dictionary of model names
        batch_num: Optional batch number (for batch saves)
        **prediction_params_kwargs: Prediction parameters
        
    Returns:
        Dictionary with file_metadata, run_metadata, and questions
    """
    # Extract model names
    model_names = get_model_names(model_dict) if model_dict else []
    
    # Count abstentions
    abstention_stats = count_abstentions(data)
    
    # Extract original metadata
    original_metadata = data.get("metadata", {})
    
    # File metadata: static info about the source file
    file_metadata = {
        "id_source": original_metadata.get("id_source"),
        "source_text": original_metadata.get("source_text"),
        "difficulty_level": original_metadata.get("difficulty_level"),
        "total_text_chunks": original_metadata.get("total_text_chunks"),
        "total_questions": original_metadata.get("total_questions"),
        "discipline": original_metadata.get("discipline")
    }
    
    # Extract prediction parameters
    prediction_params = extract_prediction_parameters(**prediction_params_kwargs)
    
    # Run metadata: dynamic info about this specific prediction run
    run_metadata = {
        "prediction_params": prediction_params if prediction_params else {},
        "models_used": model_names,
        "generated_at": datetime.now().isoformat(),
        "abstention_count": abstention_stats
    }
    
    # Add batch number if provided
    if batch_num is not None:
        run_metadata["last_batch_saved"] = batch_num
    
    return {
        "file_metadata": file_metadata,
        "run_metadata": run_metadata,
        "questions": data.get("questions", [])
    }


def _extract_prediction_text(prediction_value: Any, question_type: str) -> str:
    """
    Extract prediction text from prediction value.
    For MCQ: returns mapped_answer if available, otherwise maps raw_response using MCQ_choice_map
    For others: returns the string value directly
    
    Args:
        prediction_value: Prediction value (string or dict for MCQ)
        question_type: Question type ("MCQ", "COMP", "KNOW")
        
    Returns:
        Prediction text string
    """
    if prediction_value is None:
        return ""
    
    if normalize_question_type(question_type) == QuestionType.MCQ.value and isinstance(prediction_value, dict):
        # For MCQ, prefer mapped_answer
        mapped_answer = prediction_value.get("mapped_answer")
        if mapped_answer:
            return mapped_answer
        
        # Fallback: try to map raw_response using MCQ_choice_map_ar
        raw_response = prediction_value.get("raw_response", "")
        if raw_response:
            return MCQ_choice_map_en.get(raw_response) if raw_response in MCQ_choice_map_en else ""
        
        return ""
    elif isinstance(prediction_value, dict):
        # If it's a dict but not MCQ, try to get raw_response or convert to string
        return prediction_value.get("raw_response", str(prediction_value))
    else:
        # String value
        return str(prediction_value)


def build_question_structure(question_data: Dict[str, Any], question_type: str, q_predictions: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Build question structure with common and type-specific fields (DRY helper).
    
    Args:
        question_data: Raw question data dictionary
        question_type: Question type (MCQ, COMP, or KNOW)
        q_predictions: List of predictions for this question
        
    Returns:
        Dictionary with question structure
    """
    # Common fields for all question types
    # Handle both field name formats: "id_question" (expected) and "question_id" (actual input)
    question_id = question_data.get("id_question") or question_data.get("question_id", "")
    base_structure = {
        "id_question": question_id,
        "question": question_data.get("question", ""),
        "predictions": q_predictions
    }
    
    # Type-specific fields
    q_type = normalize_question_type(question_type)
    if q_type == QuestionType.MCQ.value:
        choices = question_data.get("choices", {})
        base_structure.update({
            "choices": choices,  # Direct dict access instead of nested .get() calls
            "correct_answer": question_data.get("correct_answer", "")
        })
    elif q_type == QuestionType.COMP.value:
        # Handle both field name formats: "id_text_chunk" (expected) and "text_chunk_id" (actual input)
        text_chunk_id = question_data.get("id_text_chunk") or question_data.get("text_chunk_id", "")
        base_structure.update({
            "id_text_chunk": text_chunk_id,
            "text": question_data.get("text", ""),
            "correct_answer": question_data.get("correct_answer", ""),
            "has_quranic_verse": question_data.get("has_quranic_verse", False),
            "has_hadith": question_data.get("has_hadith", False)
        })
    elif q_type == QuestionType.KNOW.value:
        base_structure.update({
            "correct_answer": question_data.get("correct_answer", {}),
            "has_quranic_verse": question_data.get("has_quranic_verse", False),
            "has_hadith": question_data.get("has_hadith", False),
            "no_requested_items": question_data.get("no_requested_items")
        })
    
    return base_structure


def save_batch_predictions(
    data: Dict[str, Any],
    output_path: Path,
    filename: str,
    batch_num: int,
    model_dict: Optional[Dict[str, Any]],
    question_type: str,
    save_excel: bool = True,
    **prediction_params_kwargs
) -> Dict[str, Optional[Path]]:
    """
    Save batch predictions to JSON and optionally Excel format.
    
    Updates the main output file (not a separate batch file).
    This ensures batch and final are the same file.
    
    Args:
        data: Dataset dictionary with metadata and questions (all processed so far)
        output_path: Directory to save file
        filename: Base filename (without extension)
        batch_num: Batch number (1-indexed) - used for logging only
        model_dict: Dictionary of model names
        question_type: Question type string ("MCQ", "COMP", or "KNOW")
        save_excel: Whether to also save Excel format
        **prediction_params_kwargs: Prediction parameters (few_shots, show_cot, verbalized_elicitation, etc.)
        
    Returns:
        Dictionary with paths to saved files: {"json": path, "excel": path}
    """
    saved_files = {}
    
    # Build output structure using DRY helper
    output_data = _build_output_structure(
        data, model_dict, batch_num=batch_num, **prediction_params_kwargs
    )
    
    # Save JSON to main file (not batch-numbered)
    json_file_path = output_path / "json" / f"{filename}.json"
    # Directory should already exist from config_loader.ensure_output_dirs(), but create as fallback
    json_file_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(json_file_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    saved_files["json"] = json_file_path
    print(f"  [SAVE] Batch {batch_num} saved to JSON: {json_file_path.name}")
    
    # Save Excel if requested
    if save_excel and EXCEL_AVAILABLE:
        try:
            excel_path = save_predictions_excel(data, output_path, filename, question_type, model_dict)
            saved_files["excel"] = excel_path
            if excel_path:
                print(f"  [STATS] Batch {batch_num} saved to Excel: {excel_path.name}")
        except Exception as e:
            print(f"  [WARNING] Failed to save Excel for batch {batch_num}: {e}")
            saved_files["excel"] = None
    elif save_excel and not EXCEL_AVAILABLE:
        print(f"  [WARNING] Excel export skipped for batch {batch_num} - openpyxl not installed")
        saved_files["excel"] = None
    else:
        saved_files["excel"] = None
    
    return saved_files


def save_predictions_json(
    data: Dict[str, Any],
    output_path: Path,
    filename: str,
    model_dict: Optional[Dict[str, Any]],
    **prediction_params_kwargs
) -> Path:
    """
    Save predictions to JSON format (comprehensive format).
    
    The data dictionary should already have predictions added to each question.
    
    JSON Structure (for individual files):
        {
            "file_metadata": {
                "id_source": str,
                "source_text": str,
                "difficulty_level": str,
                "total_text_chunks": int,
                "total_questions": int,
                "discipline": str
            },
            "run_metadata": {
                "prediction_params": {
                    "few_shots": bool,
                    "show_cot": bool,
                    "verbose_instructions": bool,
                    "abstention": bool,
                    "verbalized_elicitation": bool,
                    "temperature": float,
                    "max_tokens": int,
                    "word_limit": int or null,
                    "max_retries": int
                },
                "models_used": [...],
                "generated_at": "...",
                "last_batch_saved": int,
                "abstention_count": {...}
            },
            "questions": [...]
        }
    
    Args:
        data: Dataset dictionary with metadata and questions (predictions already added)
        output_path: Directory to save file
        filename: Filename base (without extension)
        model_dict: Dictionary of model names in format {"provider": ["model_name1", "model_name2", ...]}
        
    Returns:
        Path to saved file
    """
    # Build output structure using DRY helper
    output_data = _build_output_structure(data, model_dict, **prediction_params_kwargs)
    
    # Save to file
    file_path = output_path / "json" / f"{filename}.json"
    # Directory should already exist from config_loader.ensure_output_dirs(), but create as fallback
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    return file_path


def save_predictions_excel(
    data: Dict[str, Any],
    output_path: Path,
    filename: str,
    question_type: str,
    model_dict: Dict[str, Any]
) -> Optional[Path]:
    """
    Save predictions to Excel format with question-type-specific columns.
    
    Column structures:
    - MCQ: source_text, difficulty_level, question_id, question, choice1, choice2, choice3, choice4, correct_answer, model1_prediction, ...
    - COMP: source_text, difficulty_level, text_chunk_id, text_chunk, question_id, question, correct_answer, model1_prediction, ...
    - KNOW: source_text, difficulty_level, question_id, question, correct_answer, model1_prediction, ...
    
    Args:
        data: Dataset dictionary with metadata and questions (predictions already added)
        output_path: Directory to save file
        filename: Filename base (without extension)
        question_type: Question type string ("MCQ", "COMP", or "KNOW")
        model_dict: Dictionary of model names in format {"provider": ["model_name1", "model_name2", ...]}
        
    Returns:
        Path to saved file, or None if Excel not available
    """
    if not EXCEL_AVAILABLE:
        print("[WARNING] Excel export skipped - openpyxl not installed")
        return None

    # Extract models names
    model_names = get_model_names(model_dict)
    
    # Get metadata and questions
    metadata = data.get("metadata", {})
    questions = data.get("questions", [])
    
    if not questions:
        print("[WARNING] No questions found, skipping Excel export")
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    
    # Define header style (white text, bold, navy blue background)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define cell alignment (wrap text, center vertically and horizontally)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define fill colors
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for abstention
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for empty/failed
    
    # Build headers based on question type
    q_type = normalize_question_type(question_type)
    if q_type == QuestionType.MCQ.value:
        headers = [
            "Source Text", "Difficulty Level", "Question ID", "Question",
            "Choice A/أ", "Choice B/ب", "Choice C/ج", "Choice D/د", "Correct Answer"
        ]
    elif q_type == QuestionType.COMP.value:
        headers = [
            "Source Text", "Difficulty Level", "Text Chunk ID", "Text Chunk",
            "Question ID", "Question", "Correct Answer"
        ]
    else:  # KNOW
        headers = [
            "Source Text", "Difficulty Level", "Question ID", "Question", "Correct Answer", "No Requested Items"
        ]
    
    # Add model prediction columns
    for model_name in model_names:
        # Extract just the model name without provider prefix
        if "/" in model_name:
            _, model_only = model_name.split("/", 1)
        else:
            model_only = model_name
        
        if q_type == QuestionType.MCQ.value:
            # MCQ: two columns per model (mapped + raw)
            headers.append(f"{model_only}")  # Mapped response
            headers.append(f"{model_only}_raw")  # Raw response
        else:
            # KNOW/COMP: one column per model (raw only)
            headers.append(f"{model_only}_raw")  # Raw response only
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get metadata values
    source_text = metadata.get("source_text", "")
    difficulty_level = metadata.get("difficulty_level", "")
    
    # Write data rows
    row_idx = 1  # Start at 1, will increment for each valid question
    for question in questions:
        # Skip null questions
        if question is None:
            continue
        
        row_idx += 1
        col_idx = 1
        
        # Common columns: Source Text, Difficulty Level (with formatting)
        cell = ws.cell(row=row_idx, column=col_idx, value=source_text)
        cell.alignment = cell_alignment
        col_idx += 1
        cell = ws.cell(row=row_idx, column=col_idx, value=difficulty_level)
        cell.alignment = cell_alignment
        col_idx += 1
        
        # Question type specific columns (with formatting)
        if q_type == QuestionType.MCQ.value:
            # Question ID
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("id_question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Question
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Choices
            choices = question.get("choices", {})
            for choice_key in ["A", "B", "C", "D"]:
                cell = ws.cell(row=row_idx, column=col_idx, value=choices.get(choice_key, ""))
                cell.alignment = cell_alignment
                col_idx += 1
            
            # Correct Answer (show as English letter, not Arabic)
            correct_answer = question.get("correct_answer", "").strip()
            cell = ws.cell(row=row_idx, column=col_idx, value=correct_answer)
            cell.alignment = cell_alignment
            col_idx += 1
            
        elif q_type == QuestionType.COMP.value:
            # Text Chunk ID
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("id_text_chunk", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Text Chunk
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("text", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Question ID
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("id_question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Question
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Correct Answer
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("correct_answer", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
        else:  # KNOW
            # Question ID
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("id_question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Question
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Correct Answer (from correct_answer.possible_items array)
            correct_answer_obj = question.get("correct_answer", {})
            if isinstance(correct_answer_obj, dict) and "possible_items" in correct_answer_obj:
                possible_items = correct_answer_obj["possible_items"]
                # Format as semicolon-separated list for Excel
                correct_answer = " | ".join(possible_items) if possible_items else ""
            else:
                correct_answer = str(correct_answer_obj) if correct_answer_obj else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=correct_answer)
            cell.alignment = cell_alignment
            col_idx += 1
            
            # No Requested Items
            no_requested = question.get("no_requested_items")
            if no_requested is None:
                no_requested_str = "N/A"
            else:
                no_requested_str = str(no_requested)
            cell = ws.cell(row=row_idx, column=col_idx, value=no_requested_str)
            cell.alignment = cell_alignment
            col_idx += 1
        
        # Add predictions from each model
        predictions_list = question.get("predictions", [])

        # Convert list to dict: {model_name: prediction_dict}
        predictions_dict = {}
        for pred in predictions_list:
            model_name = pred.get("model", "")
            if model_name:
                predictions_dict[model_name] = pred

        for model_name in model_names:
            # Extract provider and model from "provider/model_name" format
            if "/" in model_name:
                _provider, model = model_name.split("/", 1)
            else:
                model = model_name
            
            # Look up by just model name (without provider prefix)
            prediction_value = predictions_dict.get(model, "")
            
            if q_type == QuestionType.MCQ.value:
                # MCQ: write mapped_response and raw_response (two columns)
                if prediction_value:
                    mapped_response = prediction_value.get("mapped_response", "")
                    raw_response = prediction_value.get("raw_response", "")
                else:
                    mapped_response = ""
                    raw_response = ""
                
                # Column 1: Mapped response (A/B/C/D)
                cell = ws.cell(row=row_idx, column=col_idx, value=mapped_response)
                cell.alignment = cell_alignment
                if not mapped_response:
                    cell.fill = red_fill
                col_idx += 1
                
                # Column 2: Raw response (original output)
                cell = ws.cell(row=row_idx, column=col_idx, value=raw_response)
                cell.alignment = cell_alignment
                if not raw_response:
                    cell.fill = red_fill
                elif is_abstention(raw_response):
                    cell.fill = orange_fill
                col_idx += 1
            else:
                # KNOW/COMP: write only raw_response (one column)
                raw_response = prediction_value.get("raw_response", "") if prediction_value else ""
                
                cell = ws.cell(row=row_idx, column=col_idx, value=raw_response)
                cell.alignment = cell_alignment
                
                # Apply conditional fills
                if not raw_response:
                    cell.fill = red_fill  # Red for empty/failed
                elif is_abstention(raw_response):
                    cell.fill = orange_fill  # Orange for abstention
                
                col_idx += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    # For text columns, count characters; for others, use string length
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))
            except Exception:
                pass
        # Cap width at 50, minimum 10
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save file
    file_path = output_path / "excel" / f"{filename}.xlsx"
    # Directory should already exist from config_loader.ensure_output_dirs(), but create as fallback
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(file_path))
    
    return file_path


def save_predictions(
    data: Dict[str, Any],
    output_path: Path,
    question_type: str,
    model_dict: Dict[str, Any],
    save_json: bool = True,
    save_excel: bool = True,
    filename: Optional[str] = None,
    original_file_path: Optional[Path] = None,
    **prediction_params_kwargs
) -> Dict[str, Optional[Path]]:
    """
    Save predictions in both JSON and Excel formats.
    
    Args:
        data: Dataset dictionary with metadata and questions (predictions already added)
        output_path: Directory to save files
        question_type: Question type string ("MCQ", "COMP", or "KNOW")
        model_dict: Dictionary of model names in format {"provider": ["model_name1", "model_name2", ...]}
        save_json: Whether to save JSON format (default: True)
        save_excel: Whether to save Excel format (default: True)
        filename: Optional custom filename base (without extension)
        original_file_path: Optional path to original dataset file (used for naming)
        
    Returns:
        Dictionary with paths to saved files: {"json": path, "excel": path}
    """
    saved_files = {}
    
    # Generate filename if not provided (DRY)
    if filename is None:
        filename = _generate_filename(data, original_file_path)
    
    if save_json:
        try:
            json_path = save_predictions_json(
                data, output_path, filename, model_dict,
                **prediction_params_kwargs
            )
            saved_files["json"] = json_path
        except Exception as e:
            print(f"[ERROR] Failed to save JSON: {e}")
            saved_files["json"] = None
    
    if save_excel:
        try:
            excel_path = save_predictions_excel(data, output_path, filename, question_type, model_dict)
            saved_files["excel"] = excel_path
        except Exception as e:
            print(f"[ERROR] Failed to save Excel: {e}")
            saved_files["excel"] = None
    
    return saved_files


# ========================================================================
# EVALUATION OUTPUT FUNCTIONS (for LLM-as-a-Judge)
# ========================================================================

def build_evaluation_structure(
    question_data: Dict[str, Any],
    prediction_data: Dict[str, Any],
    evaluations: List[Dict[str, Any]],
    question_type: str
) -> Dict[str, Any]:
    """
    Build evaluation structure for a single prediction with judge evaluations.
    
    **DEPRECATED**: This function builds per-prediction structures.
    Use `build_question_evaluation_structure` for the new format (per-question with predictions array).
    
    Args:
        question_data: Original question data dictionary
        prediction_data: The prediction being evaluated
        evaluations: List of evaluation results from judge models
        question_type: Question type (MCQ, COMP, or KNOW)
        
    Returns:
        Dictionary with question, single prediction, and evaluations
    """
    # Base structure with common fields
    # Handle both field name formats: "id_question" (expected) and "question_id" (actual input)
    question_id = question_data.get("id_question") or question_data.get("question_id", "")
    base_structure = {
        "id_question": question_id,
        "question": question_data.get("question", ""),
        "correct_answer": question_data.get("correct_answer", ""),
        "prediction": {
            "model": prediction_data.get("model", ""),
            "raw_response": prediction_data.get("raw_response", ""),
            "success": prediction_data.get("success", False),
            "response_time": prediction_data.get("response_time"),
            "completion_tokens": prediction_data.get("completion_tokens")
        },
        "evaluations": evaluations
    }
    
    # Add type-specific fields
    if normalize_question_type(question_type) == QuestionType.COMP.value:
        # Handle both field name formats: "id_text_chunk" (expected) and "text_chunk_id" (actual input)
        text_chunk_id = question_data.get("id_text_chunk") or question_data.get("text_chunk_id", "")
        base_structure.update({
            "id_text_chunk": text_chunk_id,
            "text": question_data.get("text", "")
        })
    
    return base_structure


def build_question_evaluation_structure(
    question_data: Dict[str, Any],
    predictions_with_evaluations: List[Dict[str, Any]],
    question_type: str
) -> Dict[str, Any]:
    """
    Build evaluation structure for a question with multiple predictions (v1 format).
    
    This matches the user's proposed structure:
    {
        "id_text_chunk": "...",  # COMP only
        "text": "...",           # COMP only
        "id_question": "...",
        "question": "...",
        "correct_answer": "...",
        "predictions": [
            {
                "model": "...",
                "raw_response": "...",
                "success": bool,
                "response_time": float,
                "completion_tokens": int,
                "evaluations": [...]
            }
        ]
    }
    
    Args:
        question_data: Original question data dictionary
        predictions_with_evaluations: List of prediction dicts, each with an "evaluations" key
        question_type: Question type (MCQ, COMP, or KNOW)
        
    Returns:
        Dictionary with question and predictions array
    """
    # Base structure with common fields
    # Handle both field name formats: "id_question" (expected) and "question_id" (actual input)
    question_id = question_data.get("id_question") or question_data.get("question_id", "")
    base_structure = {
        "id_question": question_id,
        "question": question_data.get("question", ""),
        "correct_answer": question_data.get("correct_answer", ""),
        "predictions": predictions_with_evaluations
    }
    
    # Add type-specific fields at the TOP (before id_question)
    if normalize_question_type(question_type) == QuestionType.COMP.value:
        # Handle both field name formats: "id_text_chunk" (expected) and "text_chunk_id" (actual input)
        text_chunk_id = question_data.get("id_text_chunk") or question_data.get("text_chunk_id", "")
        # Insert at the beginning for proper ordering
        ordered_structure = {
            "id_text_chunk": text_chunk_id,
            "text": question_data.get("text", "")
        }
        ordered_structure.update(base_structure)
        return ordered_structure
    
    return base_structure


def save_evaluations_json(
    data: Dict[str, Any],
    output_path: Path,
    filename: str,
    judge_dict: Dict[str, List[str]],
    original_prediction_file: Optional[str] = None,
    prediction_run_metadata: Optional[Dict[str, Any]] = None,
    **eval_params_kwargs
) -> Path:
    """
    Save evaluations to JSON format with comprehensive statistics.
    
    JSON Structure (v1 format):
        {
            "file_metadata": {
                "id_source": "...",
                "source_text": "...",
                "difficulty_level": "...",
                "total_text_chunks": N,
                "total_questions": N,
                "discipline": "..."
            },
            "prediction_metadata": {
                "prediction_params": {...},
                "models_used": [...],
                "predicted_at": "ISO-8601",
                "abstention_count": {"total": N, "per_model": {...}}
            },
            "evaluation_metadata": {
                "evaluation_params": {...},
                "judge_models": [...],
                "evaluated_at": "ISO-8601",
                "original_prediction_file": "..."
            },
            "statistics": {
                "summary": {...},
                "prediction_model_stats": {...},
                "judge_model_stats": {...}
            },
            "questions": [
                {
                    "id_text_chunk": "...",  # COMP only
                    "text": "...",           # COMP only
                    "id_question": "...",
                    "question": "...",
                    "correct_answer": "...",
                    "predictions": [
                        {
                            "model": "...",
                            "raw_response": "...",
                            "success": bool,
                            "response_time": float,
                            "completion_tokens": int,
                            "evaluations": [
                                {
                                    "judge_model": "...",
                                    "score": 0.0-1.0,
                                    "explanation": "...",
                                    "completion_tokens": int,
                                    "response_time": float,
                                    "success": bool,
                                    "generated_at": "ISO-8601"
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    
    Args:
        data: Evaluation data with questions and evaluations
        output_path: Directory to save file
        filename: Base filename (without extension)
        judge_dict: Dictionary of judge models by provider
        original_prediction_file: Name of the original prediction file
        prediction_run_metadata: Optional run_metadata from prediction file
        **eval_params_kwargs: Evaluation parameters
        
    Returns:
        Path to saved file
    """
    # Extract judge model names
    judge_models = get_model_names(judge_dict)
    
    # Extract original metadata (handle both old 'metadata' and new 'file_metadata' formats)
    original_metadata = data.get("file_metadata") or data.get("metadata", {})
    
    # File metadata: static info about the source
    file_metadata = {
        "id_source": original_metadata.get("id_source"),
        "source_text": original_metadata.get("source_text"),
        "difficulty_level": original_metadata.get("difficulty_level"),
        "total_text_chunks": original_metadata.get("total_text_chunks"),
        "total_questions": original_metadata.get("total_questions"),
        "discipline": original_metadata.get("discipline")
    }
    
    # Prediction metadata: mirrored from prediction file's run_metadata
    prediction_metadata = {}
    if prediction_run_metadata:
        prediction_metadata = {
            "prediction_params": prediction_run_metadata.get("prediction_params", {}),
            "models_used": prediction_run_metadata.get("models_used", []),
            "predicted_at": prediction_run_metadata.get("generated_at", ""),
            "abstention_count": prediction_run_metadata.get("abstention_count", {"total": 0, "per_model": {}})
        }
    
    # Extract evaluation parameters
    eval_params = extract_prediction_parameters(**eval_params_kwargs)
    
    # Evaluation metadata: info about this evaluation run
    evaluation_metadata = {
        "evaluation_params": eval_params if eval_params else {},
        "judge_models": judge_models,
        "evaluated_at": datetime.now().isoformat(),
        "original_prediction_file": original_prediction_file
    }
    
    # Calculate comprehensive statistics
    statistics = calculate_evaluation_statistics(data, prediction_metadata)
    
    output_data = {
        "file_metadata": file_metadata,
        "prediction_metadata": prediction_metadata,
        "evaluation_metadata": evaluation_metadata,
        "statistics": statistics,
        "questions": data.get("questions", [])
    }
    
    # Save to file
    file_path = output_path / "json" / f"{filename}.json"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    return file_path


def save_evaluations_excel(
    data: Dict[str, Any],
    output_path: Path,
    filename: str,
    question_type: str,
    judge_dict: Dict[str, List[str]]
) -> Optional[Path]:
    """
    Save evaluations to Excel format with judge scores.
    
    Column structures:
    - COMP: source_text, difficulty, text_chunk_id, question_id, question, text, correct_answer, 
            prediction_model, prediction, judge1_cot, judge1_correctness, judge1_relevance, 
            judge1_completeness, judge2_...
    - KNOW: source_text, difficulty, question_id, question, correct_answer, prediction_model, 
            prediction, judge1_cot, judge1_correctness, judge1_relevance, judge1_completeness, judge2_...
    
    Args:
        data: Dataset dictionary with metadata and questions (evaluations already added)
        output_path: Directory to save file
        filename: Filename base (without extension)
        question_type: Question type string ("COMP" or "KNOW")
        judge_dict: Dictionary of judge model names
        
    Returns:
        Path to saved file, or None if Excel not available
    """
    if not EXCEL_AVAILABLE:
        print("[WARNING] Excel export skipped - openpyxl not installed")
        return None
    
    # Extract judge names
    judge_names = get_model_names(judge_dict)
    
    # Get metadata and questions (handle both old 'metadata' and new 'file_metadata' formats)
    metadata = data.get("file_metadata") or data.get("metadata", {})
    questions = data.get("questions", [])
    
    if not questions:
        print("[WARNING] No questions found, skipping Excel export")
        return None
    
    # Check if any evaluation has granular scores (correctness, relevance, completeness)
    has_granular_scores = False
    for question in questions:
        if question is None:
            continue
        predictions = question.get("predictions", [])
        for prediction in predictions:
            evaluations = prediction.get("evaluations", [])
            for eval_data in evaluations:
                # Check if any granular metric exists
                if any(metric in eval_data for metric in ["correctness", "relevance", "completeness"]):
                    has_granular_scores = True
                    break
            if has_granular_scores:
                break
        if has_granular_scores:
            break
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluations"
    
    # Define header style
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define cell alignment
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Build headers based on question type
    q_type = normalize_question_type(question_type)
    if q_type == QuestionType.COMP.value:
        headers = [
            "Source Text", "Difficulty Level", "Text Chunk ID", "Question ID",
            "Question", "Text", "Correct Answer", "Prediction Model", "Confidence Level", "Prediction"
        ]
    else:  # KNOW
        headers = [
            "Source Text", "Difficulty Level", "Question ID", "Question",
            "Correct Answer", "Prediction Model", "Confidence Level", "Prediction"
        ]
    
    # Add judge columns (Score + Explanation + optional granular metrics per judge)
    for judge_name in judge_names:
        # Extract just the model name without provider prefix
        if "/" in judge_name:
            _, judge_only = judge_name.split("/", 1)
        else:
            judge_only = judge_name
        
        headers.append(f"{judge_only}_Score")
        headers.append(f"{judge_only}_Explanation")
        # Add granular metric columns only if at least one evaluation has them
        # Order: 1. Correctness, 2. Relevance, 3. Completeness
        if has_granular_scores:
            headers.append(f"{judge_only}_Correctness")
            headers.append(f"{judge_only}_Relevance")
            headers.append(f"{judge_only}_Completeness")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get metadata values
    source_text = metadata.get("source_text", "")
    difficulty_level = metadata.get("difficulty_level", "")
    
    # Write data rows - one row per prediction (new structure has predictions array)
    row_idx = 1
    for question in questions:
        if question is None:
            continue
        
        # Get predictions array (new structure)
        predictions = question.get("predictions", [])
        
        # Create one row per prediction
        for prediction in predictions:
            row_idx += 1
            col_idx = 1
            
            # Common columns
            cell = ws.cell(row=row_idx, column=col_idx, value=source_text)
            cell.alignment = cell_alignment
            col_idx += 1
            cell = ws.cell(row=row_idx, column=col_idx, value=difficulty_level)
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Question type specific columns
            if q_type == QuestionType.COMP.value:
                # Text Chunk ID
                cell = ws.cell(row=row_idx, column=col_idx, value=question.get("id_text_chunk", ""))
                cell.alignment = cell_alignment
                col_idx += 1
            
            # Question ID
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("id_question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Question
            cell = ws.cell(row=row_idx, column=col_idx, value=question.get("question", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Text (for COMP only)
            if q_type == QuestionType.COMP.value:
                cell = ws.cell(row=row_idx, column=col_idx, value=question.get("text", ""))
                cell.alignment = cell_alignment
                col_idx += 1
            
            # Correct Answer (handle both string and dict with possible_items)
            correct_answer_obj = question.get("correct_answer", "")
            if isinstance(correct_answer_obj, dict) and "possible_items" in correct_answer_obj:
                possible_items = correct_answer_obj["possible_items"]
                # Ensure all items are strings and format as pipe-separated list for Excel
                if possible_items:
                    possible_items = [str(item) for item in possible_items if item is not None]
                    correct_answer = " | ".join(possible_items) if possible_items else ""
                else:
                    correct_answer = ""
            else:
                # Convert to string, handling None and other types
                correct_answer = str(correct_answer_obj) if correct_answer_obj is not None else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=correct_answer)
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Prediction info
            cell = ws.cell(row=row_idx, column=col_idx, value=prediction.get("model", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Confidence Level
            confidence_level = prediction.get("confidence_level")
            if confidence_level is not None:
                confidence_value = f"{confidence_level}%"
            else:
                confidence_value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=confidence_value)
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Prediction (raw_response)
            cell = ws.cell(row=row_idx, column=col_idx, value=prediction.get("raw_response", ""))
            cell.alignment = cell_alignment
            col_idx += 1
            
            # Evaluations from judges (from this prediction's evaluations array)
            evaluations = prediction.get("evaluations", [])
            evaluations_dict = {eval.get("judge_model", ""): eval for eval in evaluations}
            
            for judge_name in judge_names:
                # Extract model name
                if "/" in judge_name:
                    _, judge_model = judge_name.split("/", 1)
                else:
                    judge_model = judge_name
                
                eval_data = evaluations_dict.get(judge_model, {})
                
                # Score (overall)
                score = eval_data.get("score", "")
                cell = ws.cell(row=row_idx, column=col_idx, value=score)
                cell.alignment = cell_alignment
                if score and isinstance(score, (int, float)):
                    cell.number_format = '0.0'
                col_idx += 1
                
                # Explanation (sanitize for Excel - remove illegal control characters)
                explanation = eval_data.get("explanation", "")
                if explanation:
                    # Remove control characters (except newline, tab, carriage return)
                    # Keep newlines (\n), tabs (\t), carriage returns (\r), but remove other control chars
                    explanation = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', str(explanation))
                cell = ws.cell(row=row_idx, column=col_idx, value=explanation)
                cell.alignment = cell_alignment
                col_idx += 1
                
                # Granular metrics (only write if has_granular_scores is True)
                # Order: 1. Correctness, 2. Relevance, 3. Completeness
                if has_granular_scores:
                    for metric in ["correctness", "relevance", "completeness"]:
                        value = eval_data.get(metric, "")
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = cell_alignment
                        if value and isinstance(value, (int, float)):
                            cell.number_format = '0'  # Integer format for 0-10 scores
                        col_idx += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save file
    file_path = output_path / "excel" / f"{filename}.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(file_path))
    
    print(f"  [EXCEL] Saved {len(questions)} question(s) to {file_path.name}")
    
    return file_path


def save_evaluations(
    data: Dict[str, Any],
    output_path: Path,
    question_type: str,
    judge_dict: Dict[str, List[str]],
    save_json: bool = True,
    save_excel: bool = True,
    filename: Optional[str] = None,
    original_prediction_file: Optional[str] = None,
    prediction_run_metadata: Optional[Dict[str, Any]] = None,
    **eval_params_kwargs
) -> Dict[str, Optional[Path]]:
    """
    Save evaluations in both JSON and Excel formats.
    
    Args:
        data: Dataset dictionary with metadata and questions (evaluations already added)
        output_path: Directory to save files
        question_type: Question type string ("COMP" or "KNOW")
        judge_dict: Dictionary of judge model names
        save_json: Whether to save JSON format (default: True)
        save_excel: Whether to save Excel format (default: True)
        filename: Optional custom filename base (without extension)
        original_prediction_file: Name of the original prediction file
        prediction_run_metadata: Optional run_metadata from prediction file
        **eval_params_kwargs: Evaluation parameters
        
    Returns:
        Dictionary with paths to saved files: {"json": path, "excel": path}
    """
    saved_files = {}
    
    # Generate filename if not provided
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        metadata = data.get("metadata", {})
        source_id = metadata.get("id_source", "evaluation")
        filename = f"{source_id}_{timestamp}_evaluations"
    
    if save_json:
        try:
            json_path = save_evaluations_json(
                data, output_path, filename, judge_dict,
                original_prediction_file, prediction_run_metadata, **eval_params_kwargs
            )
            saved_files["json"] = json_path
        except Exception as e:
            print(f"[ERROR] Failed to save evaluation JSON: {e}")
            saved_files["json"] = None
    
    if save_excel:
        try:
            excel_path = save_evaluations_excel(
                data, output_path, filename, question_type, judge_dict
            )
            saved_files["excel"] = excel_path
        except Exception as e:
            print(f"[ERROR] Failed to save evaluation Excel: {e}")
            print(f"[ERROR] Traceback: {traceback.format_exc()}")
            saved_files["excel"] = None
    
    return saved_files


# ========================================================================
# EVALUATION INDEX TRACKING FUNCTIONS
# ========================================================================

def _sorted_list(items) -> List[str]:
    """Convert iterable to sorted list. Compact helper."""
    return sorted(list(items))


# Note: normalize_model_name is imported from constants.py for judge name normalization


def _create_empty_question_index_entry(
    status: str, normalized_all_judges: Set[str], chunk_id: Optional[str] = None
) -> Dict[str, Any]:
    """Create a standardized empty question index entry. DRY helper."""
    entry = {
        "status": status,
        "predictions_present": [],
        "predictions_null": [],
        "evaluations_completed": [],
        "evaluations_failed": [],
        "evaluations_missing": _sorted_list(normalized_all_judges),
        "all_complete": False,
        "has_predictions": False
    }
    if chunk_id is not None:
        entry["id_text_chunk"] = chunk_id
    return entry


def track_evaluation_gaps(
    evaluation_questions: List[Dict[str, Any]],
    all_judges: Set[str],
    question_type: str,
    original_questions: Optional[List[Dict[str, Any]]] = None  # Reserved for future use
) -> Dict[str, Any]:
    """
    Track gaps (null questions, null predictions, failed/missing evaluations) for each question.
    
    Args:
        evaluation_questions: List of evaluation question dicts (from evaluation output)
        all_judges: Set of all expected judge model names (may include provider prefix)
        question_type: Question type (COMP, KNOW)
        original_questions: Optional original questions from prediction file for comparison
        
    Returns:
        Index dictionary with gap information
    """
    # Normalize all_judges to unprefixed format (matches JSON storage format)
    normalized_all_judges = {normalize_model_name(judge) for judge in all_judges}
    
    question_index = {}
    null_question_ids = []
    questions_with_null_predictions = []
    questions_with_failed_evaluations = []
    questions_with_missing_evaluations = []
    
    total_missing_evaluations = 0
    total_failed_evaluations = 0
    total_null_predictions = 0
    questions_with_gaps = 0
    
    # Process questions from evaluation file
    for q in evaluation_questions:
        # Check if null question
        if q is None:
            continue  # Skip None questions in index (they're tracked separately)
        
        q_id = q.get("id_question", "")
        if not q_id:
            continue
        
        question_text = q.get("question")
        
        # Check if null question (question text is None)
        if question_text is None:
            null_question_ids.append(q_id)
            questions_with_gaps += 1
            chunk_id = q.get("id_text_chunk") if question_type == "COMP" else None
            question_index[q_id] = _create_empty_question_index_entry(
                "null_question", normalized_all_judges, chunk_id
            )
            continue
        
        # Get predictions array
        predictions = q.get("predictions", [])
        
        # Track predictions and evaluations
        predictions_present = set()
        predictions_null = set()
        evaluations_completed = defaultdict(set)  # {model: set of judges}
        evaluations_failed = defaultdict(set)  # {model: set of judges}
        
        for pred in predictions:
            model = pred.get("model", "")
            if not model:
                continue
            
            # Check if prediction is null (success is False or raw_response is empty/None)
            if pred.get("success") is False or not pred.get("raw_response"):
                predictions_null.add(model)
                total_null_predictions += 1
            else:
                predictions_present.add(model)
            
            # Process evaluations for this prediction
            evaluations = pred.get("evaluations", [])
            for eval_data in evaluations:
                judge = eval_data.get("judge_model", "")
                if not judge:
                    continue
                
                # Normalize judge name to unprefixed format for consistent comparison
                normalized_judge = normalize_model_name(judge)
                
                if eval_data.get("success", False):
                    evaluations_completed[model].add(normalized_judge)
                else:
                    evaluations_failed[model].add(normalized_judge)
                    total_failed_evaluations += 1
        
        # Calculate missing evaluations
        # For each model with a valid prediction, check which judges are missing
        missing_evaluations = defaultdict(set)  # {model: set of missing judges}
        for model in predictions_present:
            completed_judges = evaluations_completed[model]
            failed_judges = evaluations_failed[model]
            # Use normalized_all_judges for comparison
            missing_judges = normalized_all_judges - completed_judges - failed_judges
            if missing_judges:
                missing_evaluations[model] = missing_judges
                total_missing_evaluations += len(missing_judges)
        
        # Determine status and if there are gaps
        has_gaps = False
        status = "complete"
        
        if predictions_null:
            questions_with_null_predictions.append(q_id)
            has_gaps = True
            status = "has_null_predictions"
        
        if any(evaluations_failed.values()):
            questions_with_failed_evaluations.append(q_id)
            has_gaps = True
            if status == "complete":
                status = "has_failed_evaluations"
            else:
                status = "partial"
        
        if missing_evaluations:
            questions_with_missing_evaluations.append(q_id)
            has_gaps = True
            if status == "complete":
                status = "has_missing_evaluations"
            elif status != "partial":
                status = "has_missing_evaluations"
        
        all_complete = (
            len(predictions_null) == 0 and
            not any(evaluations_failed.values()) and
            not missing_evaluations and
            len(predictions_present) > 0
        )
        
        if has_gaps or not all_complete:
            questions_with_gaps += 1
        
        # Build question index entry (ALWAYS include ALL questions - matches prediction index behavior)
        entry = {
            "status": status,
            "predictions_present": _sorted_list(predictions_present),
            "predictions_null": _sorted_list(predictions_null),
            "evaluations_completed": {
                model: _sorted_list(judges) 
                for model, judges in evaluations_completed.items()
            },
            "evaluations_failed": {
                model: _sorted_list(judges) 
                for model, judges in evaluations_failed.items()
            },
            "evaluations_missing": {
                model: _sorted_list(judges) 
                for model, judges in missing_evaluations.items()
            },
            "all_complete": all_complete,
            "has_predictions": len(predictions_present) > 0
        }
        
        if question_type == "COMP":
            chunk_id = q.get("id_text_chunk")
            if chunk_id:
                entry["id_text_chunk"] = chunk_id
        
        # CRITICAL FIX: Always include ALL questions (not just incomplete ones)
        # This ensures new judges can be detected for previously complete questions
        question_index[q_id] = entry
    
    # Calculate totals
    total_questions = len(evaluation_questions)
    valid_questions = len([q for q in evaluation_questions if q is not None and q.get("question") is not None])
    null_questions = total_questions - valid_questions
    
    # Calculate expected evaluations
    # For each valid question with valid predictions, we expect num_judges evaluations
    total_evaluations_expected = 0
    total_evaluations_completed = 0
    for q in evaluation_questions:
        if q is None or q.get("question") is None:
            continue
        predictions = q.get("predictions", [])
        for pred in predictions:
            if pred.get("success") is not False and pred.get("raw_response"):
                total_evaluations_expected += len(normalized_all_judges)
                evaluations = pred.get("evaluations", [])
                for eval_data in evaluations:
                    if eval_data.get("success", False):
                        total_evaluations_completed += 1
    
    return {
        "metadata": {
            "total_questions": total_questions,
            "valid_questions": valid_questions,
            "null_questions": null_questions,
            "questions_with_gaps": questions_with_gaps,
            "total_evaluations_expected": total_evaluations_expected,
            "total_evaluations_completed": total_evaluations_completed,
            "total_evaluations_failed": total_failed_evaluations,
            "total_evaluations_skipped": total_evaluations_expected - total_evaluations_completed - total_failed_evaluations,
            "total_predictions_null": total_null_predictions,
            "null_question_ids": sorted(null_question_ids),
            "questions_with_null_predictions": sorted(questions_with_null_predictions),
            "questions_with_failed_evaluations": sorted(questions_with_failed_evaluations),
            "questions_with_missing_evaluations": sorted(questions_with_missing_evaluations),
            "expected_judges": _sorted_list(normalized_all_judges)
        },
        "question_index": question_index
    }


def get_evaluation_index_yaml_header() -> str:
    """Get YAML file header with instructions and warnings."""
    return """# Evaluation Gap Tracking Index File
# 
# WARNING: This is an auto-generated file. Do not edit manually.
# Regenerate by re-running the evaluation pipeline.
#
# This file tracks evaluation gaps (null questions, null predictions, failed/missing evaluations).
# Use this to identify which questions need re-running for which models/judges.
#
# The script tracks:
#   - Null questions: Questions that are null in the evaluation file
#   - Null predictions: Predictions with success=false or empty raw_response
#   - Failed evaluations: Evaluations where success=false
#   - Missing evaluations: Expected evaluations that were not completed
#
# Structure:
#   - metadata: Summary statistics, expected judges, and problematic question IDs
#   - question_index: Per-question evaluation tracking (all questions included)
#     - status: "null_question", "has_null_predictions", "has_failed_evaluations", "has_missing_evaluations", "partial", or "complete"
#     - predictions_present: Models with valid predictions
#     - predictions_null: Models with null predictions
#     - evaluations_completed: {model: [judges]} - Successful evaluations per model
#     - evaluations_failed: {model: [judges]} - Failed evaluations per model
#     - evaluations_missing: {model: [judges]} - Missing evaluations per model
#     - all_complete: Whether all expected evaluations are complete
#     - id_text_chunk: For COMP questions, the text chunk ID
#
"""


def _format_evaluation_index_yaml(index_data: Dict[str, Any]) -> str:
    """
    Format evaluation index data as readable YAML with better organization and spacing.
    
    Args:
        index_data: Index data dictionary
        
    Returns:
        Formatted YAML string
    """
    lines = []
    metadata = index_data.get("metadata", {})
    question_index = index_data.get("question_index", {})
    
    # ========================================================================
    # METADATA SECTION - Organized into logical groups
    # ========================================================================
    lines.append("metadata:")
    
    # Summary Statistics
    lines.append("  # Summary Statistics")
    lines.append(f"  total_questions: {metadata.get('total_questions', 0)}")
    lines.append(f"  valid_questions: {metadata.get('valid_questions', 0)}")
    lines.append(f"  null_questions: {metadata.get('null_questions', 0)}")
    lines.append(f"  questions_with_gaps: {metadata.get('questions_with_gaps', 0)}")
    lines.append("")
    
    # Evaluation Statistics
    lines.append("  # Evaluation Statistics")
    lines.append(f"  total_evaluations_expected: {metadata.get('total_evaluations_expected', 0)}")
    lines.append(f"  total_evaluations_completed: {metadata.get('total_evaluations_completed', 0)}")
    lines.append(f"  total_evaluations_failed: {metadata.get('total_evaluations_failed', 0)}")
    lines.append(f"  total_evaluations_skipped: {metadata.get('total_evaluations_skipped', 0)}")
    lines.append("")
    
    # Prediction Statistics
    lines.append("  # Prediction Statistics")
    lines.append(f"  total_predictions_null: {metadata.get('total_predictions_null', 0)}")
    lines.append("")
    
    # Problematic Question IDs
    null_ids = metadata.get('null_question_ids', [])
    null_pred_ids = metadata.get('questions_with_null_predictions', [])
    failed_eval_ids = metadata.get('questions_with_failed_evaluations', [])
    missing_eval_ids = metadata.get('questions_with_missing_evaluations', [])
    
    lines.append("  # Problematic Question IDs")
    if null_ids:
        lines.append("  null_question_ids:")
        for q_id in null_ids:
            lines.append(f"    - {q_id}")
    else:
        lines.append("  null_question_ids: []")
    
    if null_pred_ids:
        lines.append("  questions_with_null_predictions:")
        for q_id in null_pred_ids:
            lines.append(f"    - {q_id}")
    else:
        lines.append("  questions_with_null_predictions: []")
    
    if failed_eval_ids:
        lines.append("  questions_with_failed_evaluations:")
        for q_id in failed_eval_ids:
            lines.append(f"    - {q_id}")
    else:
        lines.append("  questions_with_failed_evaluations: []")
    
    if missing_eval_ids:
        lines.append("  questions_with_missing_evaluations:")
        for q_id in missing_eval_ids:
            lines.append(f"    - {q_id}")
    else:
        lines.append("  questions_with_missing_evaluations: []")
    lines.append("")
    
    # Expected Judges
    expected_judges = metadata.get('expected_judges', [])
    if expected_judges:
        lines.append("  # Expected Judges")
        lines.append("  expected_judges:")
        for judge in expected_judges:
            lines.append(f"    - {judge}")
        lines.append("")
    
    # Index File Tracking (if present)
    evaluation_file = metadata.get('evaluation_file')
    index_generated_at = metadata.get('index_generated_at')
    index_updated_for = metadata.get('index_updated_for')
    index_regenerated = metadata.get('index_regenerated')
    
    if evaluation_file or index_generated_at or index_updated_for or index_regenerated:
        lines.append("  # Index File Tracking")
        if evaluation_file:
            lines.append(f"  evaluation_file: {evaluation_file}")
        if index_generated_at:
            lines.append(f"  index_generated_at: {index_generated_at}")
        if index_updated_for:
            lines.append(f"  index_updated_for: {index_updated_for}")
        if index_regenerated:
            lines.append(f"  index_regenerated: {index_regenerated}")
        lines.append("")
    
    # ========================================================================
    # QUESTION INDEX SECTION
    # ========================================================================
    lines.append("# Question Index")
    lines.append("# All questions are included (complete and incomplete) for comprehensive tracking")
    lines.append("question_index:")
    
    # Handle empty question_index - write explicit {} to avoid YAML parsing as None
    if not question_index:
        lines.append("  {}")
        return "\n".join(lines)
    
    # Sort questions for consistent output
    sorted_questions = sorted(question_index.items())
    
    for i, (q_id, q_data) in enumerate(sorted_questions):
        # Add spacing between questions (except first one)
        if i > 0:
            lines.append("")
        
        lines.append(f"  {q_id}:")
        
        # Status
        status = q_data.get('status')
        if status:
            lines.append(f"    status: {status}")
        
        # Predictions
        pred_present = q_data.get('predictions_present', [])
        if pred_present:
            lines.append("    predictions_present:")
            for model in sorted(pred_present):
                lines.append(f"      - {model}")
        else:
            lines.append("    predictions_present: []")
        
        pred_null = q_data.get('predictions_null', [])
        if pred_null:
            lines.append("    predictions_null:")
            for model in sorted(pred_null):
                lines.append(f"      - {model}")
        
        # Evaluations completed
        eval_completed = q_data.get('evaluations_completed', {})
        if eval_completed:
            lines.append("    evaluations_completed:")
            for model in sorted(eval_completed.keys()):
                judges = eval_completed[model]
                if judges:
                    lines.append(f"      {model}:")
                    for judge in sorted(judges):
                        lines.append(f"        - {judge}")
        
        # Evaluations failed
        eval_failed = q_data.get('evaluations_failed', {})
        if eval_failed:
            lines.append("    evaluations_failed:")
            for model in sorted(eval_failed.keys()):
                judges = eval_failed[model]
                if judges:
                    lines.append(f"      {model}:")
                    for judge in sorted(judges):
                        lines.append(f"        - {judge}")
        
        # Evaluations missing
        eval_missing = q_data.get('evaluations_missing', {})
        if eval_missing:
            lines.append("    evaluations_missing:")
            for model in sorted(eval_missing.keys()):
                judges = eval_missing[model]
                if judges:
                    lines.append(f"      {model}:")
                    for judge in sorted(judges):
                        lines.append(f"        - {judge}")
        
        # Completion status
        all_complete = q_data.get('all_complete', False)
        has_predictions = q_data.get('has_predictions', False)
        lines.append(f"    all_complete: {str(all_complete).lower()}")
        lines.append(f"    has_predictions: {str(has_predictions).lower()}")
        
        # Text chunk ID (for COMP questions)
        chunk_id = q_data.get('id_text_chunk')
        if chunk_id:
            lines.append(f"    id_text_chunk: {chunk_id}")
    
    return "\n".join(lines)


def save_evaluation_index_yaml(index_data: Dict[str, Any], output_path: Path) -> Path:
    """Save evaluation index data as YAML with header comments and improved formatting."""
    if not YAML_AVAILABLE:
        raise ImportError("PyYAML is required for YAML format. Install with: pip install pyyaml")
    
    header = get_evaluation_index_yaml_header()
    formatted_yaml = _format_evaluation_index_yaml(index_data)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(header)
        f.write(formatted_yaml)
        f.write("\n")  # Final newline
    
    return output_path



